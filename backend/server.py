from fastapi import FastAPI, APIRouter, HTTPException, Depends, Response, Request, Cookie, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from uuid import uuid4
from datetime import datetime, timezone, timedelta
import httpx
import bcrypt
from jose import JWTError, jwt
import secrets
import re
import asyncio
import tempfile
import shutil
from io import BytesIO
import jieba
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
# from multipart import parse_options_header  # Not used

# File processing imports
import PyPDF2
import openpyxl
from docx import Document
import base64
from PIL import Image

# OpenAI integration via emergentintegrations
from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# AnythingLLM API configuration (PRO version)
ANYTHINGLLM_API_KEY = os.environ.get("ANYTHINGLLM_API_KEY")
ANYTHINGLLM_BASE_URL = "https://pilj1jbx.rcsrv.com/api/v1"

# Ollama AnythingLLM API configuration (FREE version)
OLLAMA_ANYTHINGLLM_API_KEY = os.getenv('OLLAMA_API_KEY', '0PSWXGR-22AMZJP-JEEAQ1P-1EQS5DA')
OLLAMA_ANYTHINGLLM_BASE_URL = "https://2jr84ymm.rcsrv.com/api/v1"

# Keep the original API URL for backward compatibility
ANYTHINGLLM_API_URL = os.environ.get("ANYTHINGLLM_API_URL", "https://pilj1jbx.rcsrv.com/api/v1/workspace/bilgin/chat")

# Serper API configuration  
SERPER_API_KEY = os.environ.get("SERPER_API_KEY")
SERPER_API_URL = "https://google.serper.dev/search"

# OpenAI configuration via EMERGENT_LLM_KEY and direct OpenAI API
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY")
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
EMERGENT_LLM_KEY = os.environ.get("EMERGENT_LLM_KEY", "sk-emergent-42dF7720bCaB9378cD")
OPENAI_API_URL = "https://api.openai.com/v1/chat/completions"

# Gemini API configuration for FREE version
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY")

# File upload configuration
UPLOAD_DIR = Path("/tmp/bilgin_uploads")
UPLOAD_DIR.mkdir(exist_ok=True)
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# Web search functions using Serper API
async def web_search(query: str, num_results: int = 3) -> List[dict]:
    """Perform web search using Serper API"""
    if not SERPER_API_KEY:
        logging.warning("Serper API key not configured")
        return []
    
    try:
        async with httpx.AsyncClient() as client:
            payload = {
                "q": query,
                "num": min(num_results, 10),  # Serper supports up to 10 results per call
                "hl": "tr",  # Turkish language preference
                "gl": "tr"   # Turkey geographic location
            }
            
            headers = {
                "X-API-KEY": SERPER_API_KEY,
                "Content-Type": "application/json"
            }
            
            response = await client.post(
                SERPER_API_URL,
                json=payload,
                headers=headers,
                timeout=8.0
            )
            
            if response.status_code == 200:
                data = response.json()
                results = []
                
                # Process organic results
                if "organic" in data:
                    for item in data["organic"][:num_results]:
                        results.append({
                            "title": item.get("title", ""),
                            "snippet": item.get("snippet", ""),
                            "link": item.get("link", "")
                        })
                
                # Include featured snippet if available
                if "answerBox" in data and len(results) > 0:
                    answer_box = data["answerBox"]
                    if "answer" in answer_box:
                        results[0]["featured_answer"] = answer_box["answer"]
                
                logging.info(f"Serper API returned {len(results)} search results")
                return results
            else:
                logging.error(f"Serper API error: {response.status_code} - {response.text}")
                return []
                
    except Exception as e:
        logging.error(f"Web search error with Serper API: {e}")
        return []

def extract_factual_claims(text: str) -> List[str]:
    """Extract potential factual claims from AI response"""
    # Simple patterns for factual claims
    factual_patterns = [
        # "X yazmıştır" pattern
        r'([A-ZÇĞİÖŞÜ][a-zçğıöşü\s]+)\s+(?:tarafından\s+)?(?:yazılmıştır|yazmıştır|yazarı|eseri)',
        # "X yılında" pattern  
        r'(\d{4})\s*yılında\s+([^.]+)',
        # "X kişisi" pattern
        r'([A-ZÇĞİÖŞÜ][a-zçğıöşü\s]+)\s+(?:doğmuştur|ölmüştür|kurmuştur)',
        # "X şehirde" pattern
        r'([A-ZÇĞİÖŞÜ][a-zçğıöşü\s]+)\s+şehrinde',
        # Numbers and statistics
        r'(\d+(?:,\d+)*)\s+(?:metre|kilometre|kişi|yıl|gün)',
    ]
    
    claims = []
    for pattern in factual_patterns:
        matches = re.findall(pattern, text, re.IGNORECASE)
        for match in matches:
            if isinstance(match, tuple):
                claims.append(' '.join(str(m) for m in match))
            else:
                claims.append(str(match))
    
    return claims[:3]  # Limit to 3 claims to avoid too many searches

async def openai_fact_check(ai_response: str, original_query: str) -> str:
    """Use OpenAI GPT-4 to fact-check AI responses"""
    if not OPENAI_API_KEY:
        logging.warning("OpenAI API key not configured for fact-checking")
        return ai_response
    
    try:
        async with httpx.AsyncClient() as client:
            system_prompt = """Sen bir fact-checker asistanısın. Verilen cevabı doğrula ve eğer hatalı bilgi varsa düzelt. 
            Eğer cevap doğru ise olduğu gibi bırak. Eğer yanlış bilgi varsa doğru bilgiyi ver.
            Sadece kesin yanlış olan bilgileri düzelt, belirsiz durumlarda orijinal cevabı koru.
            Türkçe yanıt ver."""
            
            user_prompt = f"""Soru: {original_query}
Verilen Cevap: {ai_response}

Bu cevabı kontrol et ve eğer faktüel hata varsa düzelt. Eğer doğru ise aynen döndür."""
            
            payload = {
                "model": "gpt-4",
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ],
                "max_tokens": 600,
                "temperature": 0.1  # Lower temperature for fact-checking
            }
            
            headers = {
                "Authorization": f"Bearer {OPENAI_API_KEY}",
                "Content-Type": "application/json"
            }
            
            response = await client.post(
                OPENAI_API_URL,
                json=payload,
                headers=headers,
                timeout=15.0
            )
            
            if response.status_code == 200:
                data = response.json()
                checked_content = data["choices"][0]["message"]["content"]
                
                logging.info("OpenAI fact-checking completed successfully")
                return checked_content
            else:
                logging.error(f"OpenAI fact-checking error: {response.status_code} - {response.text}")
                return ai_response  # Return original if fact-check fails
                
    except Exception as e:
        logging.error(f"OpenAI fact-checking error: {e}")
        return ai_response  # Return original if fact-check fails

async def fact_check_with_serper(ai_response: str, original_query: str) -> str:
    """Fact-check AI response using Serper web search"""
    
    # Extract factual claims from AI response
    claims = extract_factual_claims(ai_response)
    
    if not claims:
        logging.info("No factual claims detected in AI response")
        return ai_response
    
    logging.info(f"Fact-checking with Serper: {claims}")
    
    corrections = {}
    
    for claim in claims:
        # Create search query for fact-checking
        search_query = f'"{claim}" doğru mu bilgi kontrol'
        
        # Perform web search
        search_results = await web_search(search_query, num_results=2)
        
        if search_results:
            # Simple verification logic
            verification_text = " ".join([result["snippet"] for result in search_results])
            
            # Look for contradiction indicators
            contradiction_indicators = [
                "yanlış", "hatalı", "doğru değil", "gerçek değil", 
                "aslında", "doğrusu", "gerçekte", "düzeltme"
            ]
            
            has_contradiction = any(indicator in verification_text.lower() for indicator in contradiction_indicators)
            
            if has_contradiction:
                # Try to extract correct information
                correction = await extract_correction(claim, search_results, original_query)
                if correction:
                    corrections[claim] = correction
    
    # Apply corrections to response
    corrected_response = ai_response
    for incorrect_claim, correction in corrections.items():
        if correction and correction != incorrect_claim:
            corrected_response = corrected_response.replace(incorrect_claim, correction)
            logging.info(f"Applied correction: '{incorrect_claim}' -> '{correction}'")
    
    return corrected_response

async def extract_correction(claim: str, search_results: List[dict], original_query: str) -> Optional[str]:
    """Extract corrected information from search results"""
    
    # Simple correction extraction - look for "actually" patterns
    correction_patterns = [
        r'(?:aslında|gerçekte|doğrusu)\s+([^.]+)',
        r'([A-ZÇĞİÖŞÜ][^.]+?)\s+(?:tarafından|yazmıştır)',
        r'doğru\s+(?:cevap|bilgi)\s+([^.]+)',
    ]
    
    for result in search_results:
        text = result["snippet"] + " " + result["title"]
        
        for pattern in correction_patterns:
            matches = re.findall(pattern, text, re.IGNORECASE)
            if matches:
                correction = matches[0].strip()
                if len(correction) > 5 and correction.lower() != claim.lower():
                    return correction
    
    return None

def requires_web_search(question: str) -> bool:
    """Determine if question requires web search instead of AnythingLLM"""
    
    # Güncel/live information patterns
    web_search_patterns = [
        # Sports scores and results - ENHANCED patterns
        r'(?:maç|skor|sonuç|kazandı|kaybetti)\s+(?:ne|nedir|nasıl|kaç)',
        r'(?:galatasaray|fenerbahçe|beşiktaş|trabzonspor|kasımpaşa)\s+(?:maçı|skoru|maç|sonuç)',
        r'(?:galatasaray|fenerbahçe|beşiktaş|trabzonspor)\s+(?:kasımpaşa|galatasaray|fenerbahçe|beşiktaş)',
        r'(?:barcelona|real madrid|manchester|liverpool)\s+(?:maç|skor)',
        r'(?:şampiyonlar ligi|premier lig|süper lig)\s+(?:sonuç|tablo)',
        r'maç\s+sonuç',  # Direct "maç sonuç" pattern
        r'(?:son|güncel|bugünkü)\s+maç',  # "son maç" pattern
        
        # Current news and events  
        r'(?:son|güncel|bugünkü|şu an)\s+(?:haber|durum|gelişme)',
        r'(?:bugün|dün|geçen hafta)\s+(?:ne oldu|olan)',
        r'(?:seçim|politika|hükümet)\s+(?:son|güncel)',
        
        # Financial data
        r'(?:dolar|euro|bitcoin|borsa)\s+(?:kuru|fiyat|ne kadar)',
        r'(?:altın|gümüş|petrol)\s+(?:fiyat|kur)',
        
        # Weather
        r'(?:hava durumu|hava|yağmur|kar)\s+(?:nasıl|ne|bugün)',
        r'(?:ankara|istanbul|izmir)\s+(?:hava|sıcaklık)',
        
        # Recent releases/publications
        r'(?:yeni|son)\s+(?:çıkan|yayınlanan)\s+(?:kitap|film|müzik)',
        r'(?:2024|2025)\s+(?:çıkan|yayınlanan)',
        
        # Books/works not likely in system
        r'(?:hangi|kim)\s+(?:yazdı|yazarı|eseri)',  # Could be unknown works
        r'(?:kitap|roman|şiir)\s+(?:kim|kimin|hangi)',
        
        # Live data
        r'(?:şu an|anlık|canlı)\s+',
        r'(?:en son|en yeni|fresh)',
        
        # Technology and recent events
        r'(?:chatgpt|ai|yapay zeka)\s+(?:son|yeni|güncel)',
        r'(?:iphone|android|tesla)\s+(?:yeni|son|model)'
    ]
    
    question_lower = question.lower()
    
    # Debug logging
    for i, pattern in enumerate(web_search_patterns):
        if re.search(pattern, question_lower):
            logging.info(f"Web search triggered by pattern {i}: '{pattern}' for question: '{question}'")
            return True
    
    logging.info(f"No web search pattern matched for question: '{question}'")
    return False

async def clean_web_search_with_anythingllm(web_search_result: str, original_question: str) -> str:
    """Clean and improve web search results using AnythingLLM - REMOVE source attribution"""
    
    try:
        # Create a cleaning prompt for AnythingLLM - specifically ask to remove source info
        cleaning_prompt = f"""Lütfen aşağıdaki web araştırması sonucunu düzenle ve daha okunabilir hale getir.
        
Orijinal Soru: {original_question}

Web Araştırması Sonucu:
{web_search_result}

Lütfen bu bilgiyi:
1. Daha açık ve anlaşılır hale getir
2. Gereksiz tekrarları kaldır  
3. Türkçe dilbilgisi kurallarına uygun düzenle
4. Ana bilgileri öne çıkar
5. "Web araştırması sonucunda", "güncel web kaynaklarından" gibi kaynak belirtimlerini KALDIR
6. Sadece temiz, doğrudan cevap ver

Temiz cevap:"""

        async with httpx.AsyncClient() as client:
            api_payload = {
                "message": cleaning_prompt,
                "mode": "chat",
                "sessionId": f"cleaning-{hash(original_question)}"
            }
            
            response = await client.post(
                ANYTHINGLLM_API_URL,
                headers={
                    "Authorization": f"Bearer {ANYTHINGLLM_API_KEY}",
                    "Content-Type": "application/json"
                },
                json=api_payload,
                timeout=20.0
            )
            
            if response.status_code == 200:
                ai_response = response.json()
                cleaned_result = ai_response.get("textResponse", web_search_result)
                
                # Fix English error messages from AnythingLLM
                if "sorry, i'm experiencing technical difficulties" in cleaned_result.lower():
                    cleaned_result = "Üzgünüm, şu anda teknik bir sorun yaşıyorum. Lütfen sorunuzu tekrar deneyin."
                elif "sorry" in cleaned_result.lower() and "technical" in cleaned_result.lower():
                    cleaned_result = "Teknik sorun nedeniyle temizleme yapılamadı. Orijinal web sonucu gösteriliyor."
                
                # Additional cleaning - remove any remaining source attributions
                cleaned_result = re.sub(r'\*.*web.*kaynak.*\*', '', cleaned_result, flags=re.IGNORECASE)
                cleaned_result = re.sub(r'web araştırması sonucunda:?', '', cleaned_result, flags=re.IGNORECASE)
                cleaned_result = re.sub(r'güncel.*kaynak.*alınmıştır', '', cleaned_result, flags=re.IGNORECASE)
                cleaned_result = cleaned_result.strip()
                
                logging.info("Web search result cleaned and source attribution removed")
                return cleaned_result
            else:
                logging.error(f"AnythingLLM cleaning error: {response.status_code}")
                return web_search_result
                
    except Exception as e:
        logging.error(f"Web search cleaning error: {e}")
        return web_search_result

async def get_anythingllm_response(question: str, conversation_mode: str = 'normal') -> str:
    """Get response from AnythingLLM"""
    
    try:
        # Apply conversation mode prompts if needed
        final_message = question
        if conversation_mode and conversation_mode != 'normal':
            mode_prompts = {
                'friend': "Lütfen samimi, motive edici ve esprili bir şekilde yanıtla. 3 küçük adım önerebilirsin. Arkadaş canlısı ol:",
                'realistic': "Eleştirel ve kanıt odaklı düşün. Güçlü ve zayıf yönleri belirt. Test planı öner. Gerçekci ol:",
                'coach': "Soru sorarak kullanıcının düşünmesini sağla. Hedef ve adım listesi çıkar. Koç gibi yaklaş:",
                'lawyer': "Bilinçli karşı argüman üret. Kör noktaları göster. Avukat perspektifiyle yaklaş:",
                'teacher': "Adım adım öğret. Örnek ver ve mini quiz ekle. Öğretmen gibi açıkla:",
                'minimalist': "En kısa, madde işaretli, süssüz yanıt ver. Minimalist ol:"
            }
            if conversation_mode in mode_prompts:
                final_message = f"{mode_prompts[conversation_mode]} {question}"
        
        async with httpx.AsyncClient() as client:
            api_payload = {
                "message": final_message,
                "mode": "chat",
                "sessionId": f"hybrid-{hash(question)}"
            }
            
            response = await client.post(
                ANYTHINGLLM_API_URL,
                headers={
                    "Authorization": f"Bearer {ANYTHINGLLM_API_KEY}",
                    "Content-Type": "application/json"
                },
                json=api_payload,
                timeout=20.0
            )
            
            if response.status_code == 200:
                ai_response = response.json()
                raw_response = ai_response.get("textResponse", "AnythingLLM yanıt veremedi.")
                
                # Fix common English error messages - MORE COMPREHENSIVE
                response_lower = raw_response.lower().strip()
                
                # Check for exact English error message match
                if response_lower == "sorry, i'm experiencing technical difficulties.":
                    return "Üzgünüm, şu anda teknik bir sorun yaşıyorum. Lütfen sorunuzu tekrar deneyin."
                elif "sorry, i'm experiencing technical difficulties" in response_lower:
                    return "Üzgünüm, şu anda teknik bir sorun yaşıyorum. Lütfen sorunuzu tekrar deneyin."
                elif "sorry" in response_lower and "experiencing" in response_lower and "technical" in response_lower:
                    return "Teknik bir sorun nedeniyle yanıt veremedim. Lütfen tekrar deneyin."
                elif "sorry" in response_lower and ("technical" in response_lower or "difficulties" in response_lower):
                    return "Üzgünüm, teknik sorun yaşıyorum. Lütfen tekrar deneyin."
                elif "i cannot" in response_lower or "i can't" in response_lower:
                    return "Bu konuda yardımcı olamıyorum. Başka bir şey sorabilirsiniz."
                elif response_lower.startswith("sorry"):
                    return "Üzgünüm, o konuda yardımcı olamıyorum."
                
                # Clean markdown formatting from response
                cleaned_response = clean_response_formatting(raw_response)
                return cleaned_response
            else:
                logging.error(f"AnythingLLM error: {response.status_code}")
                return "AnythingLLM'e şu anda erişilemedi. Lütfen tekrar deneyin."
                
    except Exception as e:
        logging.error(f"AnythingLLM request error: {e}")
        return "AnythingLLM ile bağlantı sorunu yaşandı. Lütfen tekrar deneyin."

async def handle_web_search_question(question: str) -> str:
    """Handle questions that require web search using Serper API"""
    
    # Perform web search with optimized query
    search_query = optimize_search_query(question)
    search_results = await web_search(search_query, num_results=3)
    
    if not search_results:
        return "Üzgünüm, şu an web araması yapamıyorum. Lütfen daha sonra tekrar deneyin."
    
    # Format web search results into natural response
    response = format_web_search_response(question, search_results)
    return response

def can_anythingllm_answer(anythingllm_response: str) -> bool:
    """Check if AnythingLLM provided answer - specifically look for 'NO_ANSWER' pattern"""
    
    response_stripped = anythingllm_response.strip()
    
    # Check for the specific "NO_ANSWER\nSources:" pattern from AnythingLLM
    if "NO_ANSWER" in response_stripped and "Sources:" in response_stripped:
        logging.info("AnythingLLM returned 'NO_ANSWER\\nSources:' - RAG system has no information")
        return False
    
    # Check for other no answer patterns
    response_lower = response_stripped.lower()
    
    if "no answer" in response_lower:
        logging.info("AnythingLLM returned 'no answer' - RAG system has no information")
        return False
    
    # Check for very short responses (likely inadequate)
    if len(response_stripped) < 10:
        logging.info("Response too short - considering as inadequate")
        return False
    
    # Check for clear inability indicators
    no_answer_indicators = [
        "bilmiyorum", "hiç bilmiyorum", "bu konuda bilgim yok",
        "hiçbir fikrim yok", "cevap veremiyorum", "bilgi bulamıyorum",
        "erişemiyorum", "bağlantı hatası", "yanıt veremedi",
        "sorry", "i don't know", "can't access", "unable to"
    ]
    
    for indicator in no_answer_indicators:
        if indicator in response_lower:
            logging.info(f"AnythingLLM cannot answer: '{indicator}' found in response")
            return False
    
    # If AnythingLLM provided a substantive response, use it
    logging.info("AnythingLLM provided answer from RAG system - using it")
    return True

def get_question_category(question: str) -> str:
    """Quickly categorize question type for optimal routing"""
    
    question_lower = question.lower().strip()
    
    # Simple greetings and casual questions - AnythingLLM only
    casual_patterns = [
        r'^(merhaba|selam|naber|nasılsın|iyi misin)$',
        r'^(hello|hi|hey)$',
        r'(teşekkür|sağol|eyvallah)',
        r'^(tamam|ok|peki|anladım)$',
        r'(nasıl gidiyor|keyifler|ne yapıyorsun)',
        r'^(iyi geceler|günaydın|tünaydın)$'
    ]
    
    # Current/live information - Web search required (Google'dan aratılabilecek bilgiler)
    current_info_patterns = [
        # Time-sensitive information
        r'(bugün|şu an|anlık|canlı|güncel|son|şimdi)\s+',
        r'(dün|geçen hafta|bu hafta|son hafta|yeni)\s+',
        
        # Weather - Google'dan aratılabilir
        r'(hava durumu|hava|sıcaklık|yağmur|kar|rüzgar|nem)',
        r'(meteoroloji|hava tahmini|iklim)',
        
        # Sports - Google'dan aratılabilir  
        r'(maç|skor|sonuç|puan|derbi|lig)\s*',
        r'(galatasaray|fenerbahçe|beşiktaş|trabzonspor).*(maç|skor)',
        r'(barcelona|real madrid|manchester|chelsea).*(maç|skor)',
        r'(şampiyonlar ligi|premier lig|süper lig|la liga).*(sonuç|tablo|puan)',
        r'(futbol|basketbol|voleybol).*(sonuç|maç)',
        
        # Financial/Currency - Google'dan aratılabilir
        r'(dolar|euro|bitcoin|altın|borsa|kripto)\s+(kur|fiyat|değer)',
        r'(bist|nasdaq|dow jones|piyasa)',
        
        # News and Events - Google'dan aratılabilir
        r'(haber|gelişme|olay|açıklama).*(son|yeni|bugün|güncel)',
        r'(seçim|politika|hükümet|cumhurbaşkanı).*(son|güncel)',
        r'(deprem|afet|kaza|olay).*(son|bugün)',
        
        # Traffic and Transportation - Google'dan aratılabilir
        r'(trafik|yol durumu|ulaşım|metro|otobüs)',
        r'(kapalı|açık).*(yol|köprü|tünel)',
        
        # Recent releases/publications - Google'dan aratılabilir
        r'(yeni|son).*(çıkan|yayınlanan|piyasaya).*(kitap|film|müzik|oyun)',
        r'(2024|2025).*(çıkan|yayınlanan|çıkacak)',
        
        # Store hours, opening times - Google'dan aratılabilir
        r'(açık|kapalı|saat).*(market|mağaza|restoran|banka)',
        r'(çalışma saatleri|açılış saati)',
        
        # Live events, concerts - Google'dan aratılabilir
        r'(konser|etkinlik|festival|gösteri).*(bugün|yakında|tarih)',
        
        # Current prices, availability - Google'dan aratılabilir
        r'(fiyat|ücret|maliyet).*(şu an|güncel|bugün)',
        r'(satış|indirim|kampanya).*(aktif|geçerli)'
    ]
    
    # Check for casual questions first
    for pattern in casual_patterns:
        if re.search(pattern, question_lower):
            return 'casual'
    
    # Check for current information needs (Google'dan aratılabilecek bilgiler)
    for pattern in current_info_patterns:
        if re.search(pattern, question_lower):
            return 'current'
    
    # Check if it's a factual knowledge question that might need verification
    factual_patterns = [
        r'(kim|hangi|ne zaman|nerede).*(yazdı|yaptı|oldu|kurdu)',
        r'(kaç|ne kadar).*(yıl|metre|kilo|kişi)',
        r'(başkenti|nüfusu|yüzölçümü)',
        r'(doğum|ölüm).*(tarih|yıl)',
        r'(eseri|kitabı|şiiri|filmi)'
    ]
    
    for pattern in factual_patterns:
        if re.search(pattern, question_lower):
            return 'factual'
    
    # Math or calculation questions - AnythingLLM preferred
    if re.search(r'(\d+\s*[\+\-\*\/x×÷]\s*\d+|matematik|hesap|kaç\s+eder)', question_lower):
        return 'math'
    
    # Default to general knowledge
    return 'general'

def are_responses_similar(response1: str, response2: str) -> bool:
    """Check if two responses contain similar information to avoid duplication"""
    
    if not response1 or not response2:
        return False
    
    # Remove common prefixes/suffixes for comparison
    clean1 = re.sub(r'(web araştırması sonucunda:?|bilgin cevabı:?)', '', response1.lower())
    clean2 = re.sub(r'(web araştırması sonucunda:?|bilgin cevabı:?)', '', response2.lower())
    
    # Simple similarity check - if one response contains most words from the other
    words1 = set(clean1.split())
    words2 = set(clean2.split())
    
    if not words1 or not words2:
        return False
    
    # Remove common short words
    common_words = {'bir', 'bu', 'şu', 'da', 'de', 've', 'ile', 'için', 'olan', 'the', 'and', 'or', 'in', 'on', 'at'}
    words1 = words1 - common_words
    words2 = words2 - common_words
    
    if not words1 or not words2:
        return False
    
    # Calculate overlap
    intersection = len(words1.intersection(words2))
    smaller_set_size = min(len(words1), len(words2))
    
    # If 60% or more words overlap, consider similar
    similarity = intersection / smaller_set_size
    
    logging.info(f"Response similarity: {similarity:.2f} (threshold: 0.6)")
    return similarity >= 0.6

async def process_with_openai_gpt5_nano(question: str, conversation_mode: str = 'normal', file_content: str = None, file_name: str = None) -> str:
    """Process question with ChatGPT-4o-mini using Emergent integrations"""
    try:
        # Prepare personality prompt if conversation mode is active
        if conversation_mode and conversation_mode != 'normal':
            mode_personalities = {
                'friend': "Sanki karşında bir arkadaşın varmış gibi davranır, sohbeti takip eder, dostça fikirler ve samimi yorumlar yapar. Arkadaşça, sıcak ve samimi bir dille konuş.",
                'realistic': "Yalan söylemez, yanlış fikri açıkça belirtir, seni en doğru sonuca ulaştırmaya çalışır. Gerçekçi, objektif ve doğrudan yaklaşım sergile.",
                'coach': "Sorular sorarak düşündürür, hedef belirlemene yardımcı olur, adım adım ilerleme planı çıkarır. Motivasyonel koç gibi davran ve sürekli sorular sor.",
                'lawyer': "Karşıt görüş üretir, bir avukat gibi mantıklı ve savunmacı şekilde karşı argümanı savunur. Her konuda alternatif bakış açısı sun ve karşı argüman geliştir.",
                'teacher': "Her konuda öğretici yaklaşır, konuyu adım adım açıklar ve pekiştirme soruları sorarak anlamanı kontrol eder. Pedagogik ve eğitici bir dil kullan.",
                'minimalist': "Tek cümlelik, net ve doğrudan cevap verir; uzatmaz, sadece soruna odaklanır. Çok kısa ve öz yanıtlar ver, gereksiz ayrıntıya girme."
            }
            system_message = mode_personalities.get(conversation_mode, "Sen Türkçe konuşan yardımcı bir asistansın.")
        else:
            system_message = "Sen Türkçe konuşan yardımcı bir asistansın. Kullanıcının sorularına sadece doğru, güvenilir ve kanıta dayalı cevaplar verirsin. Emin olmadığın konularda 'emin değilim' dersin ve yanlış bilgi vermezsin."
        
        # Initialize chat with Emergent LLM
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"pro-{uuid.uuid4()}",
            system_message=system_message
        ).with_model("openai", "gpt-4o-mini")
        
        # Prepare the message
        if file_content:
            user_text = f"Dosya adı: {file_name}\n\nDosya içeriği:\n{file_content}\n\nKullanıcı sorusu: {question}\n\nLütfen bu dosya hakkındaki soruyu yanıtla."
        else:
            user_text = question
        
        user_message = UserMessage(text=user_text)
        
        # Send message and get response
        response = await chat.send_message(user_message)
        
        # Clean markdown formatting from response
        cleaned_response = clean_response_formatting(response)
        logging.info("ChatGPT-4o-mini response received successfully via Emergent integrations")
        return cleaned_response
                
    except Exception as e:
        logging.error(f"ChatGPT-4o-mini request error via Emergent integrations: {e}")
        return "ChatGPT API'sine bağlanırken bir hata oluştu. Lütfen tekrar deneyin."

def is_formula_based_question(question: str) -> bool:
    """Check if question requires formula calculations or specialized technical knowledge"""
    formula_keywords = [
        # Matematik formülleri
        'integral', 'türev', 'diferansiyel', 'eşitlik', 'denklem', 'formula', 'hesapla', 'çöz',
        'sin', 'cos', 'tan', 'logaritma', 'üssel', 'kök', 'karekök', 'faktöriyel',
        'matris', 'determinant', 'vektör', 'trigonometri',
        # Fizik formülleri  
        'newton', 'euler', 'maxwell', 'schrödinger', 'ohm', 'coulomb', 'bernoulli',
        'termodinamik', 'kinetik', 'potansiyel', 'momentum', 'enerji korunumu',
        # Mühendislik formülleri
        'mukavemet', 'gerilme', 'burulma', 'moment', 'kiriş', 'yapı analizi',
        'elektrik devre', 'impedans', 'frekans', 'filtre tasarımı',
        # İstatistik formülleri
        'standart sapma', 'varyans', 'korelasyon', 'regresyon', 'hipotez testi',
        'normal dağılım', 'p-değeri', 'güven aralığı'
    ]
    
    question_lower = question.lower()
    return any(keyword in question_lower for keyword in formula_keywords)

def is_general_knowledge_question(question: str) -> bool:
    """Check if question is about general knowledge, culture, history, etc."""
    general_keywords = [
        # Genel kültür
        'tarih', 'coğrafya', 'sanat', 'müzik', 'edebiyat', 'sinema', 'spor',
        'ünlü kişi', 'ünlü', 'meşhur', 'kim', 'nerede', 'hangi yıl', 'ne zaman',
        'başkent', 'ülke', 'şehir', 'dil', 'kültür', 'din', 'bayram', 'festival',
        # Yaşam tarzı
        'yemek', 'tarif', 'seyahat', 'gezi', 'tatil', 'hobi', 'eğlence',
        'sağlık', 'beslenme', 'egzersiz', 'moda', 'stil', 'ev dekorasyonu',
        # İnsan bilimi
        'psikoloji', 'sosyoloji', 'felsefe', 'ahlak', 'etik', 'yaşam', 'mutluluk',
        'başarı', 'kariyer', 'ilişki', 'aile', 'çocuk', 'eğitim'
    ]
    
    question_lower = question.lower()
    return any(keyword in question_lower for keyword in general_keywords)

async def simple_pro_system(question: str, conversation_mode: str = 'normal', file_content: str = None, file_name: str = None) -> str:
    """PRO system with smart routing: ChatGPT for general knowledge, AnythingLLM for formula-based questions"""
    
    logging.info(f"PRO version - Smart routing system for: {question}")
    
    # Step 1: Check if conversation mode is active - use Ollama for all conversation modes
    if conversation_mode and conversation_mode != 'normal':
        logging.info(f"PRO: Conversation mode {conversation_mode} detected - using Ollama AnythingLLM")
        return await process_with_ollama_free(question, conversation_mode, file_content, file_name)
    
    # Step 2: Check if question is about current/güncel topics
    category = get_question_category(question)
    if category == 'current':
        logging.info("PRO: Current topic detected - using web search")
        web_search_response = await handle_web_search_question(question)
        return await clean_web_search_with_anythingllm(web_search_response, question)
    
    # Step 3: Smart routing for normal questions
    if is_formula_based_question(question):
        logging.info("PRO: Formula-based question detected - using AnythingLLM for specialized knowledge")
        try:
            anythingllm_response = await get_anythingllm_response(question, conversation_mode)
            if can_anythingllm_answer(anythingllm_response):
                return anythingllm_response
            else:
                logging.info("PRO: AnythingLLM couldn't answer formula question - falling back to ChatGPT")
                return await process_with_openai_gpt5_nano(question, conversation_mode, file_content, file_name)
        except Exception as e:
            logging.error(f"PRO: AnythingLLM error for formula question: {e} - using ChatGPT")
            return await process_with_openai_gpt5_nano(question, conversation_mode, file_content, file_name)
    elif is_general_knowledge_question(question):
        logging.info("PRO: General knowledge question detected - using ChatGPT directly")
        return await process_with_openai_gpt5_nano(question, conversation_mode, file_content, file_name)
    else:
        logging.info("PRO: Standard question - trying AnythingLLM first")
        try:
            anythingllm_response = await get_anythingllm_response(question, conversation_mode)
            
            # Check if AnythingLLM gave "no answer" or error
            if can_anythingllm_answer(anythingllm_response):
                logging.info("PRO: AnythingLLM provided good answer - using it")
                return anythingllm_response
            else:
                logging.info("PRO: AnythingLLM returned 'no answer' or error - falling back to ChatGPT")
                return await process_with_openai_gpt5_nano(question, conversation_mode, file_content, file_name)
                
        except Exception as e:
            logging.error(f"PRO: AnythingLLM error: {e} - falling back to ChatGPT")
            return await process_with_openai_gpt5_nano(question, conversation_mode, file_content, file_name)

def optimize_search_query(question: str) -> str:
    """Optimize question for better web search results"""
    
    # Remove question words and optimize for search
    question_clean = question.lower()
    
    # Remove Turkish question words
    remove_words = [
        'nasıl', 'ne', 'nedir', 'kim', 'nerede', 'ne zaman', 'kaç', 'hangi',
        'mıdır', 'midir', 'mudur', 'müdür', 'mi', 'mı', 'mu', 'mü'
    ]
    
    for word in remove_words:
        question_clean = re.sub(f'\\b{word}\\b', '', question_clean)
    
    # Clean up extra spaces
    question_clean = ' '.join(question_clean.split())
    
    # Add context for better results
    if 'maç' in question_clean or 'skor' in question_clean:
        question_clean += ' sonuç skor'
    elif 'kitap' in question_clean or 'roman' in question_clean:
        question_clean += ' yazar eser bilgi'
    elif 'haber' in question_clean:
        question_clean += ' güncel haberler'
    
    return question_clean

def format_web_search_response(question: str, search_results: List[dict]) -> str:
    """Format web search results into natural conversational response"""
    
    if not search_results:
        return "Bu konuda güncel bilgi bulamadım."
    
    # Combine search results into coherent response
    main_info = []
    for result in search_results:
        snippet = result.get("snippet", "").strip()
        if snippet and len(snippet) > 10:
            main_info.append(snippet)
    
    if not main_info:
        return "Bu konuda detaylı bilgi bulamadım."
    
    # Create natural response
    response = "Web araştırması sonucunda:\n\n"
    
    # Use first result as main answer
    response += main_info[0]
    
    # Add supplementary info if available
    if len(main_info) > 1:
        response += f"\n\nEk bilgi: {main_info[1]}"
    
    # Add source indicator
    response += f"\n\n*Güncel web kaynaklarından alınmıştır ({len(search_results)} sonuç)*"
    
    return response

async def extract_text_from_file(file_path: str, file_type: str) -> str:
    """Extract text from various file types"""
    try:
        if file_type == 'pdf':
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text.strip()
        
        elif file_type in ['xlsx', 'xls']:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n--- {sheet_name} ---\n"
                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
            return text.strip()
        
        elif file_type == 'docx':
            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        
        elif file_type == 'txt':
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read().strip()
        
        elif file_type in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp']:
            # For images, we don't extract text here - we'll use Vision API
            return f"[GÖRSEL DOSYASI: {file_path}]"
        
        else:
            return "Desteklenmeyen dosya türü."
            
    except Exception as e:
        logging.error(f"File text extraction error: {e}")
        return f"Dosya okuma hatası: {str(e)}"

def get_file_type(filename: str) -> str:
    """Get file type from filename"""
    extension = filename.lower().split('.')[-1]
    return extension

def encode_image_to_base64(image_path: str) -> str:
    """Convert image to base64 for ChatGPT Vision API"""
    with open(image_path, "rb") as image_file:
        return base64.b64encode(image_file.read()).decode('utf-8')

async def search_web_for_free_version(question: str) -> str:
    """Search web using Serper API for FREE version current information"""
    try:
        headers = {
            "X-API-KEY": SERPER_API_KEY,
            "Content-Type": "application/json"
        }
        
        payload = {
            "q": question,
            "gl": "tr",  # Turkey
            "hl": "tr",  # Turkish
            "num": 5
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(SERPER_API_URL, headers=headers, json=payload, timeout=10.0)
            
            if response.status_code == 200:
                data = response.json()
                
                # Extract results
                results = []
                if 'organic' in data:
                    for item in data['organic'][:3]:  # Top 3 results
                        title = item.get('title', '')
                        snippet = item.get('snippet', '')
                        if title and snippet:
                            results.append(f"• {title}: {snippet}")
                
                if results:
                    web_content = "\n".join(results)
                    logging.info("Serper API search successful for FREE version")
                    return web_content
                else:
                    return "Web araması sonuç bulamadı."
                    
            else:
                logging.error(f"Serper API error: {response.status_code}")
                return "Web araması sırasında hata oluştu."
                
    except Exception as e:
        logging.error(f"Serper web search error: {e}")
        return "Web araması sırasında hata oluştu."

async def clean_web_results_with_gemini(web_results: str, question: str, conversation_mode: str = 'normal') -> str:
    """Clean and process web search results using Gemini API"""
    try:
        headers = {
            "Content-Type": "application/json"
        }
        
        # Prepare personality prompt if needed
        if conversation_mode and conversation_mode != 'normal':
            mode_personalities = {
                'friend': "Sanki karşında bir arkadaşın varmış gibi davranır, sohbeti takip eder, dostça fikirler ve samimi yorumlar yapar.",
                'realistic': "Yalan söylemez, yanlış fikri açıkça belirtir, seni en doğru sonuca ulaştırmaya çalışır.",
                'coach': "Sorular sorarak düşündürür, hedef belirlemene yardımcı olur, adım adım ilerleme planı çıkarır.",
                'lawyer': "Karşıt görüş üretir, bir avukat gibi mantıklı ve savunmacı şekilde karşı argümanı savunur.",
                'teacher': "Her konuda öğretici yaklaşır, konuyu adım adım açıklar ve pekiştirme soruları sorarak anlamanı kontrol eder.",
                'minimalist': "Tek cümlelik, net ve doğrudan cevap verir; uzatmaz, sadece soruna odaklanır."
            }
            personality = mode_personalities.get(conversation_mode, "Sen yardımcı bir asistansın.")
        else:
            personality = "Sen yardımcı bir asistansın."
        
        cleaning_prompt = f"""Sistem: {personality} Aşağıdaki web araması sonuçlarını kullanarak kullanıcının sorusunu Türkçe olarak yanıtla. Sadece önemli ve doğru bilgileri kullan, gereksiz detayları çıkar.

Web Araması Sonuçları:
{web_results}

Kullanıcı Sorusu: {question}

Lütfen bu bilgileri temizleyip düzenli bir Türkçe cevap ver. Kaynakları belirtme, sadece temiz ve anlaşılır bilgiyi sun."""
        
        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": cleaning_prompt
                        }
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 1,
                "maxOutputTokens": 1000
            }
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}",
                headers=headers,
                json=payload,
                timeout=30.0
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get('candidates') and data['candidates'][0].get('content'):
                    content = data['candidates'][0]['content']['parts'][0]['text']
                    logging.info("Gemini web result cleaning successful")
                    return content
                else:
                    return web_results  # Fallback to original web results
            else:
                logging.error(f"Gemini cleaning error: {response.status_code}")
                return web_results  # Fallback to original web results
                
    except Exception as e:
        logging.error(f"Gemini cleaning error: {e}")
        return web_results  # Fallback to original web results

def clean_response_formatting(text: str) -> str:
    """Clean markdown formatting from AI responses"""
    if not text:
        return text
    
    # Remove bold formatting (**text**)
    text = re.sub(r'\*\*([^*]+)\*\*', r'\1', text)
    
    # Remove italic formatting (*text*)
    text = re.sub(r'\*([^*]+)\*', r'\1', text)
    
    # Remove header formatting (### text)
    text = re.sub(r'^###\s*(.+)$', r'\1', text, flags=re.MULTILINE)
    text = re.sub(r'^##\s*(.+)$', r'\1', text, flags=re.MULTILINE)
    text = re.sub(r'^#\s*(.+)$', r'\1', text, flags=re.MULTILINE)
    
    return text

async def process_with_ollama_free(question: str, conversation_mode: str = 'normal', file_content: str = None, file_name: str = None) -> str:
    """Process question with Ollama AnythingLLM for FREE/PRO version - returns exact response without modification"""
    try:
        # Use Ollama AnythingLLM API
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {OLLAMA_ANYTHINGLLM_API_KEY}"
        }
        
        # Add conversation mode personalities
        mode_personalities = {
            'friend': "Sanki karşında bir arkadaşın varmış gibi davranır, sohbeti takip eder, dostça fikirler ve samimi yorumlar yapar. Arkadaşça, sıcak ve samimi bir dille konuş.",
            'realistic': "Yalan söylemez, yanlış fikri açıkça belirtir, seni en doğru sonuca ulaştırmaya çalışır. Gerçekçi, objektif ve doğrudan yaklaşım sergile.",
            'coach': "Sorular sorarak düşündürür, hedef belirlemene yardımcı olur, adım adım ilerleme planı çıkarır. Motivasyonel koç gibi davran ve sürekli sorular sor.",
            'lawyer': "Karşıt görüş üretir, bir avukat gibi mantıklı ve savunmacı şekilde karşı argümanı savunur. Her konuda alternatif bakış açısı sun ve karşı argüman geliştir.",
            'teacher': "Her konuda öğretici yaklaşır, konuyu adım adım açıklar ve pekiştirme soruları sorarak anlamanı kontrol eder. Pedagogik ve eğitici bir dil kullan.",
            'minimalist': "Tek cümlelik, net ve doğrudan cevap verir; uzatmaz, sadece soruna odaklanır. Çok kısa ve öz yanıtlar ver, gereksiz ayrıntıya girme."
        }
        
        # Prepare the message with personality if conversation mode is active
        if conversation_mode and conversation_mode != 'normal':
            personality = mode_personalities.get(conversation_mode, "Sen yardımcı bir asistansın.")
            if file_content:
                user_message = f"Sistem mesajı: {personality}\n\nDosya adı: {file_name}\nDosya içeriği: {file_content}\n\nKullanıcı sorusu: {question}"
            else:
                user_message = f"Sistem mesajı: {personality}\n\nKullanıcı sorusu: {question}"
        else:
            # Normal mode
            if file_content:
                user_message = f"Dosya adı: {file_name}\nDosya içeriği: {file_content}\n\nKullanıcı sorusu: {question}"
            else:
                user_message = question
        
        payload = {
            "message": user_message,
            "mode": "chat",  # Use chat mode for general knowledge with custom embeddings
            "sessionId": f"free-session-{hash(question)}",
            "reset": False
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(
                f"{OLLAMA_ANYTHINGLLM_BASE_URL}/workspace/testtt/chat",
                headers=headers,
                json=payload,
                timeout=30.0
            )
            
            if response.status_code == 200:
                data = response.json()
                # Return the exact response without any modification
                if data.get('textResponse'):
                    content = data['textResponse']
                    # Clean markdown formatting from response
                    content = clean_response_formatting(content)
                    logging.info("Ollama AnythingLLM FREE response received successfully")
                    return content
                else:
                    logging.error(f"Ollama AnythingLLM unexpected response format: {data}")
                    return "Yanıt alınamadı. Lütfen tekrar deneyin."
            else:
                logging.error(f"Ollama AnythingLLM API error: {response.status_code} - {response.text}")
                return "AnythingLLM sisteminde bir hata oluştu. Lütfen tekrar deneyin."
                
    except Exception as e:
        logging.error(f"Ollama AnythingLLM request error: {e}")
        return "AnythingLLM sistemine bağlanırken bir hata oluştu. Lütfen tekrar deneyin."

async def process_with_gemini_free(question: str, conversation_mode: str = 'normal', file_content: str = None, file_name: str = None) -> str:
    """Process question with free Gemini API for FREE version - includes web search for current topics"""
    try:
        # Check if this is a current information question that needs web search
        category = get_question_category(question)
        
        if category == 'current':
            logging.info("FREE version: Current information question detected - using Serper + Gemini")
            # Get web search results first
            web_results = await search_web_for_free_version(question)
            
            if web_results and "hata" not in web_results.lower():
                # Clean web results with Gemini
                return await clean_web_results_with_gemini(web_results, question, conversation_mode)
            else:
                # Fallback to regular Gemini if web search fails
                logging.info("Web search failed, falling back to regular Gemini")
        
        # Regular Gemini processing for non-current questions or web search fallback
        logging.info("FREE version: Using regular Gemini API")
        
        # Headers for Gemini API
        headers = {
            "Content-Type": "application/json"
        }
        
        # Prepare the message based on conversation mode
        if conversation_mode and conversation_mode != 'normal':
            mode_personalities = {
                'friend': "Sanki karşında bir arkadaşın varmış gibi davranır, sohbeti takip eder, dostça fikirler ve samimi yorumlar yapar. Arkadaşça, sıcak ve samimi bir dille konuş.",
                'realistic': "Yalan söylemez, yanlış fikri açıkça belirtir, seni en doğru sonuca ulaştırmaya çalışır. Gerçekçi, objektif ve doğrudan yaklaşım sergile.",
                'coach': "Sorular sorarak düşündürür, hedef belirlemene yardımcı olur, adım adım ilerleme planı çıkarır. Motivasyonel koç gibi davran ve sürekli sorular sor.",
                'lawyer': "Karşıt görüş üretir, bir avukat gibi mantıklı ve savunmacı şekilde karşı argümanı savunur. Her konuda alternatif bakış açısı sun ve karşı argüman geliştir.",
                'teacher': "Her konuda öğretici yaklaşır, konuyu adım adım açıklar ve pekiştirme soruları sorarak anlamanı kontrol eder. Pedagogik ve eğitici bir dil kullan.",
                'minimalist': "Tek cümlelik, net ve doğrudan cevap verir; uzatmaz, sadece soruna odaklanır. Çok kısa ve öz yanıtlar ver, gereksiz ayrıntıya girme."
            }
            system_prompt = mode_personalities.get(conversation_mode, "Sen yardımsever bir asistansın.")
        else:
            system_prompt = "Sen Türkçe konuşan yardımcı bir asistansın."
        
        # Prepare the content
        if file_content:
            user_content = f"Sistem: {system_prompt}\n\nDosya adı: {file_name}\nDosya içeriği: {file_content}\n\nKullanıcı sorusu: {question}"
        else:
            user_content = f"Sistem: {system_prompt}\n\nKullanıcı sorusu: {question}"
        
        payload = {
            "contents": [
                {
                    "parts": [
                        {
                            "text": user_content
                        }
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 1,
                "maxOutputTokens": 1500
            }
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_API_KEY}",
                headers=headers,
                json=payload,
                timeout=30.0
            )
            
            if response.status_code == 200:
                data = response.json()
                if data.get('candidates') and data['candidates'][0].get('content'):
                    content = data['candidates'][0]['content']['parts'][0]['text']
                    logging.info("Gemini FREE API response received successfully")
                    return content
                else:
                    logging.error(f"Gemini API unexpected response format: {data}")
                    return "Gemini API'sinde beklenmeyen yanıt formatı. Lütfen tekrar deneyin."
            else:
                logging.error(f"Gemini API error: {response.status_code} - {response.text}")
                return "Gemini API'sinde bir hata oluştu. Lütfen tekrar deneyin."
                
    except Exception as e:
        logging.error(f"Gemini FREE system error: {e}")
        return "Gemini FREE sisteminde bir hata oluştu. Lütfen tekrar deneyin."

def get_image_mime_type(file_path: str) -> str:
    """Detect image mime type from file extension"""
    extension = file_path.lower().split('.')[-1]
    mime_mapping = {
        'jpg': 'image/jpeg',
        'jpeg': 'image/jpeg',
        'png': 'image/png',
        'gif': 'image/gif',
        'bmp': 'image/bmp',
        'webp': 'image/webp'
    }
    return mime_mapping.get(extension, 'image/jpeg')

async def process_image_with_chatgpt_vision(question: str, image_path: str, image_name: str) -> str:
    """Process image questions using Emergent integrations text-only (temporarily disabled Vision)"""
    try:
        logging.info(f"Processing image question (text-only mode): {question}")
        
        # For now, we'll process image questions without actually seeing the image
        # This is a temporary solution until Vision API key issue is resolved
        
        # Initialize chat with Emergent LLM (text-only)
        try:
            chat = LlmChat(
                api_key=EMERGENT_LLM_KEY,
                session_id=f"image-text-{uuid.uuid4()}",
                system_message="Sen bir görsel analiz asistanısın. Kullanıcı resim yüklemiş ama şu anda görsel analiz yapamıyorsun. Resmin içeriğini tahmin etmek yerine, kullanıcıya resmi tanımlamasını rica et ve genel bilgiler ver."
            ).with_model("openai", "gpt-4o-mini")
            
            # Create text-only user message
            user_message = UserMessage(
                text=f"Kullanıcı bir resim yükledi ({image_name}) ve şu soruyu soruyor: '{question}'\n\nŞu anda görsel analiz yapamıyorum. Lütfen kullanıcıya resmi kısaca tanımlamasını rica et ve sorusuna genel bir yaklaşımla yardım et."
            )
            
            # Send message and get response
            response = await chat.send_message(user_message)
            
            # Clean response
            cleaned_response = clean_response_formatting(response)
            logging.info(f"Text-only image response generated successfully: {len(cleaned_response)} characters")
            return cleaned_response
            
        except Exception as emergent_error:
            logging.error(f"Emergent integrations error: {emergent_error}")
            
            # Fallback to informative message
            return f"Resim '{image_name}' yüklenmiş ve sorunuz: '{question}'\n\nŞu anda görsel analiz özelliği geçici olarak kullanılamıyor. Lütfen resminizi kısaca tanımlayın, size o konuda yardımcı olabilirim. Örneğin resimde matematik problemi, grafik, şema veya başka bir içerik varsa bunları yazılı olarak paylaşabilirsiniz."
                
    except Exception as e:
        logging.error(f"Image processing error: {e}")
        return f"Resim işleme sırasında bir hata oluştu. Lütfen resminizi tanımlayın, size yardımcı olabilirim."

def is_file_processing_question(question: str) -> bool:
    """Check if question requires file processing capabilities"""
    file_processing_keywords = [
        'pdf', 'excel', 'word', 'dosya', 'döküman', 'belge',
        'özet', 'özetle', 'çevir', 'translate', 'düzelt', 'analiz',
        'inceleme', 'rapor', 'tablo', 'grafik', 'veri',
        'metin düzelt', 'grammar', 'yazım', 'imla'
    ]
    
    question_lower = question.lower()
    return any(keyword in question_lower for keyword in file_processing_keywords)

def is_casual_chat(question: str) -> bool:
    """Check if this is casual conversational text"""
    casual_patterns = [
        # Greetings
        r'^(merhaba|selam|hello|hi|hey|günaydın|iyi akşam|tünaydın)$',
        r'^(merhaba|selam|hello|hi|hey)\s*(nasılsın|naber|ne var ne yok).*',
        
        # Social conversation
        r'nasılsın', r'naber', r'ne yapıyorsun', r'keyifler nasıl',
        r'ne var ne yok', r'hayat nasıl', r'işler nasıl',
        
        # Casual responses
        r'^(teşekkür|sağol|eyvallah|thanks|thank you)$',
        r'^(tamam|ok|peki|anladım|iyi|güzel)$',
        
        # Conversation starters
        r'sohbet etmek', r'konuşmak istiyorum', r'muhabbet',
        r'canım sıkılıyor', r'vakit geçirmek', r'chat yapalım',
        
        # Personal sharing
        r'bugün.*oldu', r'dün.*gittim', r'şimdi.*yapıyorum',
        r'hissediyorum', r'mutluyum', r'üzgünüm', r'yorgunum',
        
        # Simple questions that invite conversation
        r'^(nasıl|ne|neden).*\?$'
    ]
    
    question_lower = question.lower().strip()
    
    # Very short conversational expressions
    if len(question_lower) <= 15 and any(word in question_lower for word in ['merhaba', 'selam', 'naber', 'nasıl', 'hi', 'hello']):
        return True
    
    for pattern in casual_patterns:
        if re.search(pattern, question_lower):
            return True
    
    return False

def is_technical_or_creative_question(question: str) -> bool:
    """Check if question requires technical writing or creative capabilities"""
    technical_creative_keywords = [
        # Writing and content creation
        'yaz', 'yazı yaz', 'metin yaz', 'makale yaz', 'blog yaz',
        'hikaye yaz', 'şiir yaz', 'mektup yaz', 'email yaz',
        'rapor yaz', 'özgeçmiş yaz', 'sunum hazırla',
        
        # Document processing and summarization  
        'özetle', 'özet çıkar', 'kısalt', 'ana nokta', 'önemli kısmı',
        'özet ver', 'özetini al', 'özetle',
        
        # Text correction and improvement
        'düzelt', 'yazım hatası', 'imla hatası', 'grammar', 
        'dil bilgisi', 'yazım kontrolü', 'metni düzelt',
        'daha iyi yaz', 'geliştirebilir misin', 'iyileştir',
        
        # Translation
        'çevir', 'translate', 'tercüme', 'çevirisi', 'ingilizcesi',
        'türkçesi', 'fransızcası', 'almancası', 'İngilizce çevir',
        
        # Creative tasks
        'tasarla', 'fikir ver', 'öneri', 'senaryo yaz', 'plan yap',
        'strateji geliş', 'yaratıcı', 'kreatif', 'konsept',
        
        # Technical analysis
        'analiz et', 'değerlendir', 'incele', 'yorumla',
        'karşılaştır', 'araştır', 'detaylı', 'derinlemesine',
        
        # Professional tasks
        'iş planı', 'proje planı', 'sunum', 'toplantı notları',
        'agenda', 'şablon', 'format', 'profesyonel'
    ]
    
    question_lower = question.lower()
    return any(keyword in question_lower for keyword in technical_creative_keywords)

async def process_conversation_mode_with_openai(question: str, conversation_mode: str, file_content: str = None, file_name: str = None) -> str:
    """Process conversation modes with OpenAI GPT-5-nano for better personality"""
    try:
        # Define detailed personalities for each conversation mode
        mode_personalities = {
            'friend': {
                'name': 'Arkadaş Canlısı',
                'system_message': """Sanki karşında bir arkadaşın varmış gibi davranır, sohbeti takip eder, dostça fikirler ve samimi yorumlar yapar. Arkadaşça, sıcak ve samimi bir dille konuş. Empati kurarsın ve kullanıcının yanında olduğunu hissettirirsin. Dostane tavırla yaklaşır, sohbeti sürdürecek yorumlar yaparsın."""
            },
            'realistic': {
                'name': 'Gerçekçi',
                'system_message': """Yalan söylemez, yanlış fikri açıkça belirtir, seni en doğru sonuca ulaştırmaya çalışır. Gerçekçi, objektif ve doğrudan yaklaşım sergile. Güçlü ve zayıf yönleri dengeli şekilde belirtirsin. Romantik hayaller yerine gerçek durumu açık şekilde anlat."""
            },
            'coach': {
                'name': 'Koç',
                'system_message': """Sorular sorarak düşündürür, hedef belirlemene yardımcı olur, adım adım ilerleme planı çıkarır. Motivasyonel koç gibi davran ve sürekli sorular sor. Kullanıcının kendi cevaplarını bulmasına yardım edersin. Hedef belirleme ve aksiyona odaklanırsın."""
            },
            'lawyer': {
                'name': 'Avukat',
                'system_message': """Karşıt görüş üretir, bir avukat gibi mantıklı ve savunmacı şekilde karşı argümanı savunur. Her konuda alternatif bakış açısı sun ve karşı argüman geliştir. Risk analizini çok iyi yaparsın. Detayları gözden kaçırmazsın. Kanıt odaklı ve sistematik yaklaşırsın."""
            },
            'teacher': {
                'name': 'Öğretmen',
                'system_message': """Her konuda öğretici yaklaşır, konuyu adım adım açıklar ve pekiştirme soruları sorarak anlamanı kontrol eder. Pedagogik ve eğitici bir dil kullan. Karmaşık konuları basit şekilde açıklarsın. Adım adım öğretirsin ve örnekler verirsin. Mini quiz veya pratik sorular sorarsın."""
            },
            'minimalist': {
                'name': 'Minimalist',
                'system_message': """Tek cümlelik, net ve doğrudan cevap verir; uzatmaz, sadece soruna odaklanır. Çok kısa ve öz yanıtlar ver, gereksiz ayrıntıya girme. Madde işaretli ve net yapıda cevaplar verirsin. Süssüz, direkt ve işlevsel yaklaşırsın."""
            }
        }
        
        # Get the personality for the current mode
        personality = mode_personalities.get(conversation_mode, mode_personalities['friend'])
        
        # Use direct OpenAI API with the provided key
        headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json"
        }
        
        # Prepare the message
        if file_content:
            user_message = f"Dosya adı: {file_name}\n\nDosya içeriği:\n{file_content}\n\nKullanıcı sorusu: {question}\n\nLütfen bu dosya hakkındaki soruyu kişiliğine uygun şekilde yanıtla."
        else:
            user_message = question
        
        payload = {
            "model": "gpt-4o-mini",
            "messages": [
                {"role": "system", "content": personality['system_message']},
                {"role": "user", "content": user_message}
            ],
            "max_tokens": 1000,
            "temperature": 0.3  # Moderate temperature for personality
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=30.0
            )
            
            if response.status_code == 200:
                data = response.json()
                content = data["choices"][0]["message"]["content"]
                
                # GPT-5-nano sometimes returns empty content, check and handle
                if not content or content.strip() == "":
                    logging.warning("GPT-5-nano returned empty content in conversation mode")
                    content = "Üzgünüm, yanıt üretilirken bir sorun oluştu. Lütfen sorunuzu farklı şekilde tekrar deneyin."
                
                logging.info(f"OpenAI conversation mode '{conversation_mode}' response received successfully")
                return content
            else:
                logging.error(f"OpenAI conversation mode API error: {response.status_code} - {response.text}")
                return f"OpenAI API'sinde bir hata oluştu. {personality['name']} modunda yanıt veremedim."
                
    except Exception as e:
        logging.error(f"OpenAI conversation mode request error: {e}")
        return f"OpenAI API'sine bağlanırken bir hata oluştu. Lütfen tekrar deneyin."

async def process_with_direct_openai(question: str, file_content: str = None, file_name: str = None) -> str:
    """Process question with OpenAI GPT-5-nano for technical/creative tasks"""
    try:
        # Use direct OpenAI API with the provided key
        headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json"
        }
        
        # Prepare the message
        if file_content:
            user_message = f"Dosya adı: {file_name}\n\nDosya içeriği:\n{file_content}\n\nKullanıcı isteği: {question}\n\nLütfen bu dosya hakkındaki isteği profesyonel bir şekilde yerine getir."
        else:
            user_message = question
        
        system_message = "Sen profesyonel bir yazım asistanı, editör ve içerik üreticisisin. Metin yazma, düzeltme, çeviri, özet çıkarma ve yaratıcı içerik üretme konularında uzmansın. Her zaman kaliteli, doğru ve kullanıcı dostu yanıtlar verirsin. Yanlış bilgi vermemek için dikkatli olur ve emin olmadığın konularda bunu belirtirsin."
        
        payload = {
            "model": "gpt-4o-mini",
            "messages": [
                {"role": "system", "content": system_message},
                {"role": "user", "content": user_message}
            ],
            "max_tokens": 1000,
            "temperature": 0.3
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=30.0
            )
            
            if response.status_code == 200:
                data = response.json()
                content = data["choices"][0]["message"]["content"]
                
                # GPT-5-nano sometimes returns empty content, check and handle
                if not content or content.strip() == "":
                    logging.warning("GPT-5-nano returned empty content in direct API")
                    content = "Üzgünüm, yanıt üretilirken bir sorun oluştu. Lütfen sorunuzu farklı şekilde tekrar deneyin."
                
                logging.info("Direct OpenAI API response received successfully")
                return content
            else:
                logging.error(f"Direct OpenAI API error: {response.status_code} - {response.text}")
                return "OpenAI API'sinde bir hata oluştu. Lütfen tekrar deneyin."
                
    except Exception as e:
        logging.error(f"Direct OpenAI API request error: {e}")
        return "OpenAI API'sine bağlanırken bir hata oluştu. Lütfen tekrar deneyin."

def is_question_about_uploaded_file(question: str) -> bool:
    """Check if the question is specifically about an uploaded file"""
    file_reference_keywords = [
        'pdf', 'dosya', 'döküman', 'belge', 'excel', 'word',
        'yüklediğim', 'bu dosya', 'bu pdf', 'bu belge',
        'dosyada', 'pdf\'de', 'belgede', 'tabloda',
        'içerik', 'metin', 'veri', 'bilgi',
        'özet', 'özetle', 'çevir', 'analiz', 'inceleme',
        'bu', 'şu', 'o' # simple referrals when context has files
    ]
    
    question_lower = question.lower()
    
    # Direct file references
    direct_references = [
        'pdf', 'dosya', 'döküman', 'belge', 'excel', 'word',
        'yüklediğim', 'bu dosya', 'bu pdf', 'bu belge'
    ]
    
    # Check for direct file references first
    for keyword in direct_references:
        if keyword in question_lower:
            return True
    
    # Check for file processing actions with context words
    processing_actions = ['özet', 'özetle', 'çevir', 'analiz', 'inceleme', 'düzelt']
    context_words = ['bu', 'şu', 'o', 'içerik', 'metin', 'veri']
    
    has_processing_action = any(action in question_lower for action in processing_actions)
    has_context_word = any(context in question_lower for context in context_words)
    
    if has_processing_action and has_context_word:
        return True
    
    return False

async def process_with_openai(question: str, file_content: str = None, file_name: str = None) -> str:
    """Process question with OpenAI GPT-4o mini"""
    try:
        # Initialize LLM chat with EMERGENT_LLM_KEY
        chat = LlmChat(
            api_key=EMERGENT_LLM_KEY,
            session_id=f"file-processing-{hash(question)}",
            system_message="Sen Türkçe konuşan yardımcı bir asistansın. Dosya analizi, özet çıkarma, çeviri, metin düzeltme ve Excel işleme konularında uzmansın."
        ).with_model("openai", "gpt-5-nano")
        
        # Prepare the message
        if file_content:
            user_message = UserMessage(
                text=f"Dosya adı: {file_name}\n\nDosya içeriği:\n{file_content}\n\nKullanıcı sorusu: {question}\n\nLütfen bu dosya hakkındaki soruyu yanıtla."
            )
        else:
            user_message = UserMessage(text=question)
        
        # Send message to OpenAI
        response = await chat.send_message(user_message)
        
        logging.info("OpenAI GPT-4o mini response received successfully")
        return response
        
    except Exception as e:
        logging.error(f"OpenAI processing error: {e}")
        return f"OpenAI işleme hatası: {str(e)}"

def should_fact_check(ai_response: str) -> bool:
    """Determine if AI response should be fact-checked"""
    
    # Skip fact-checking for certain response types
    skip_patterns = [
        r'^(?:merhaba|selam|nasılsın)',  # Greetings
        r'(?:teşekkür|sağol|eyvallah)',  # Thanks
        r'(?:anlamadım|anlayamadım)',    # Confusion
        r'(?:üzgünüm|maalesef)',         # Apologies
        r'^\d+[\+\-\*\/]\d+',           # Simple math
    ]
    
    for pattern in skip_patterns:
        if re.search(pattern, ai_response.lower()):
            return False
    
    # Fact-check if contains potential factual claims
    factual_indicators = [
        r'tarafından\s+yazılmıştır',  # Written by
        r'\d{4}\s*yılında',          # In year
        r'[A-ZÇĞİÖŞÜ][a-zçğıöşü\s]+\s+(?:doğmuş|ölmüş|kurmuş)',  # Born/died
        r'\d+\s*(?:metre|kilometre|kişi)',  # Numbers with units
        r'başkenti|nüfusu|yüzölçümü'  # Geographic facts
    ]
    
    return any(re.search(pattern, ai_response, re.IGNORECASE) for pattern in factual_indicators)

# JWT Configuration
SECRET_KEY = os.environ.get("JWT_SECRET_KEY", secrets.token_urlsafe(32))
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 7

# Security
security = HTTPBearer(auto_error=False)

# Admin credentials
ADMIN_USERNAME = "kaantas77"
ADMIN_PASSWORD_HASH = bcrypt.hashpw("64Ekremimaro2004MKüge.".encode('utf-8'), bcrypt.gensalt())

# Define Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: EmailStr
    password_hash: str
    name: Optional[str] = None  # User's display name
    is_verified: bool = False
    is_admin: bool = False
    onboarding_completed: bool = False  # Track if user completed onboarding
    oauth_provider: Optional[str] = None
    oauth_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    last_login: Optional[datetime] = None

class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str

class UserLogin(BaseModel):
    username: str
    password: str

class OnboardingData(BaseModel):
    name: str

class UserResponse(BaseModel):
    id: str
    username: str
    email: str
    name: Optional[str] = None
    is_verified: bool
    is_admin: bool
    onboarding_completed: bool = False
    created_at: str
    last_login: Optional[str] = None

class ReportCreate(BaseModel):
    message: str
    user_agent: Optional[str] = None
    url: Optional[str] = None

class Report(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid4()))
    user_id: str
    message: str
    user_agent: Optional[str] = None
    url: Optional[str] = None
    status: str = "open"  # open, investigating, resolved
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    
class ReportResponse(BaseModel):
    id: str
    user_id: str
    message: str
    user_agent: Optional[str] = None
    url: Optional[str] = None
    status: str
    created_at: str

class Session(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    session_token: str
    expires_at: datetime
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class Conversation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    title: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ConversationCreate(BaseModel):
    title: str

class Message(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    conversation_id: str
    role: str  # "user" or "assistant"
    content: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MessageCreate(BaseModel):
    content: str
    mode: str = "chat"  # "chat" or "query" 
    conversationMode: Optional[str] = None  # Test için konuşma modları
    version: str = "pro"  # "pro" or "free" - version selection

class FileUpload(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    conversation_id: str
    file_name: str
    file_type: str
    file_path: str
    uploaded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class MessageResponse(BaseModel):
    id: str
    conversation_id: str
    role: str
    content: str
    timestamp: datetime

class AdminStats(BaseModel):
    total_users: int
    verified_users: int
    total_conversations: int
    total_messages: int
    recent_users: List[UserResponse]

# Helper functions
def prepare_for_mongo(data):
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, datetime):
                data[key] = value.isoformat()
    return data

def parse_from_mongo(item):
    if isinstance(item, dict):
        for key, value in item.items():
            # Only convert datetime strings for models that expect datetime objects
            # UserResponse expects strings, so we skip conversion for user objects
            if key in ['expires_at'] and isinstance(value, str):
                try:
                    item[key] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                except:
                    pass
    return item

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, password_hash: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), password_hash.encode('utf-8'))

def create_access_token(data: dict):
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt, expire

async def get_current_user(request: Request, session_token: Optional[str] = Cookie(None)) -> Optional[dict]:
    """Get current user from session token (cookie or header)"""
    token = session_token
    
    # Fallback to Authorization header if no cookie
    if not token:
        auth_header = request.headers.get("Authorization")
        if auth_header and auth_header.startswith("Bearer "):
            token = auth_header.split(" ")[1]
    
    if not token:
        return None
    
    try:
        # Check if token exists in database and is not expired
        session = await db.sessions.find_one({
            "session_token": token,
            "expires_at": {"$gt": datetime.now(timezone.utc).isoformat()}
        })
        
        if not session:
            return None
        
        # Get user
        user = await db.users.find_one({"id": session["user_id"]})
        if user:
            return parse_from_mongo(user)
        
    except Exception as e:
        logging.error(f"Error getting current user: {e}")
        return None
    
    return None

async def require_auth(request: Request, session_token: Optional[str] = Cookie(None)) -> dict:
    """Require authentication"""
    user = await get_current_user(request, session_token)
    if not user:
        raise HTTPException(status_code=401, detail="Not authenticated")
    return user

async def require_admin(request: Request, session_token: Optional[str] = Cookie(None)) -> dict:
    """Require admin privileges"""
    user = await require_auth(request, session_token)
    if not user.get("is_admin"):
        raise HTTPException(status_code=403, detail="Admin privileges required")
    return user

def generate_conversation_title(message_content: str) -> str:
    """Generate a short and meaningful conversation title based on the first message content"""
    message = message_content.strip()
    
    # Simple conversational messages should be labeled as "Sohbet"
    casual_greetings = ["merhaba", "selam", "naber", "nasılsın", "hey", "hi", "hello", "iyi misin", "sa", "selamlar"]
    if any(greeting in message.lower() for greeting in casual_greetings) and len(message.split()) <= 3:
        return "Sohbet"
    
    # Kısa ve anlamlı başlık oluşturma
    words = message.split()
    
    # Soruları tespit et ve konu çıkar
    if "karşılaştır" in message.lower():
        # "Yalova mı büyük avcılar mı" -> "Yalova Avcılar karşılaştırması"
        meaningful_words = [w for w in words if len(w) > 2 and w.lower() not in ["mı", "mi", "mu", "mü", "büyük", "küçük", "iyi", "kötü", "nasıl", "nerede", "ne", "kim"]]
        if len(meaningful_words) >= 2:
            return f"{meaningful_words[0]} {meaningful_words[1]} karşılaştırması"
    
    if any(q in message.lower() for q in ["nedir", "ne demek", "açıkla"]):
        # "Python nedir" -> "Python"
        meaningful_words = [w for w in words if len(w) > 2 and w.lower() not in ["nedir", "ne", "demek", "açıkla", "olan", "için", "hakkında"]]
        if meaningful_words:
            return meaningful_words[0]
    
    if "nasıl" in message.lower():
        # "Python nasıl öğrenilir" -> "Python öğrenme"
        meaningful_words = [w for w in words if len(w) > 2 and w.lower() not in ["nasıl", "yapılır", "yaparım", "öğrenilir", "için", "olan"]]
        if len(meaningful_words) >= 1:
            return f"{meaningful_words[0]} öğrenme" if len(meaningful_words) == 1 else f"{meaningful_words[0]} {meaningful_words[1]}"
    
    # Genel durumlar için ilk anlamlı kelimeleri al
    meaningful_words = [w for w in words if len(w) > 2 and w.lower() not in ["bir", "bu", "şu", "o", "ile", "için", "olan", "gibi", "kadar", "hakkında", "ve", "veya"]]
    
    if len(meaningful_words) >= 2:
        return f"{meaningful_words[0]} {meaningful_words[1]}"
    elif len(meaningful_words) == 1:
        return meaningful_words[0]
    
    # Fallback: Genel sohbet mesajı
    return "Sohbet"

# Initialize admin user
async def init_admin():
    """Initialize admin user and update existing users"""
    # Update existing users with missing fields
    await db.users.update_many(
        {"name": {"$exists": False}},
        {"$set": {"name": None, "onboarding_completed": False}}
    )
    
    # Create admin user if not exists
    admin_user = await db.users.find_one({"username": ADMIN_USERNAME})
    if not admin_user:
        admin = User(
            username=ADMIN_USERNAME,
            email="admin@bilgin.ai",
            password_hash=ADMIN_PASSWORD_HASH.decode('utf-8'),
            name="Admin",
            is_verified=True,
            is_admin=True,
            onboarding_completed=True
        )
        admin_dict = prepare_for_mongo(admin.dict())
        await db.users.insert_one(admin_dict)
        logging.info("Admin user created")
    else:
        # Update admin with missing fields if needed
        await db.users.update_one(
            {"username": ADMIN_USERNAME},
            {"$set": {
                "name": "Admin",
                "onboarding_completed": True
            }}
        )

# Authentication Routes
@api_router.post("/auth/register")
async def register_user(user_data: UserCreate):
    # Check if user already exists
    existing_user = await db.users.find_one({
        "$or": [
            {"email": user_data.email},
            {"username": user_data.username}
        ]
    })
    
    if existing_user:
        raise HTTPException(status_code=400, detail="User already exists")
    
    # Create new user
    user = User(
        username=user_data.username,
        email=user_data.email,
        password_hash=hash_password(user_data.password)
    )
    
    user_dict = prepare_for_mongo(user.dict())
    await db.users.insert_one(user_dict)
    
    return {"message": "User registered successfully. Please verify your email."}

@api_router.post("/auth/login")
async def login_user(user_data: UserLogin, response: Response):
    try:
        logging.info(f"Login attempt for user: {user_data.username}")
        
        # Find user by username or email
        user = await db.users.find_one({
            "$or": [
                {"username": user_data.username},
                {"email": user_data.username}  # Allow login with email in username field
            ]
        })
        
        if not user:
            logging.warning(f"User not found: {user_data.username}")
            raise HTTPException(status_code=400, detail="Invalid credentials")
            
        if not verify_password(user_data.password, user["password_hash"]):
            logging.warning(f"Invalid password for user: {user_data.username}")
            raise HTTPException(status_code=400, detail="Invalid credentials")
        
        user = parse_from_mongo(user)
        logging.info(f"User login successful: {user['username']}")
        
        # Create session token
        session_token, expires_at = create_access_token({"user_id": user["id"]})
        
        # Save session to database
        session = Session(
            user_id=user["id"],
            session_token=session_token,
            expires_at=expires_at
        )
        session_dict = prepare_for_mongo(session.dict())
        await db.sessions.insert_one(session_dict)
        
        # Update last login
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Set cookie
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            path="/",
            expires=expires_at
        )
        
        user_response = UserResponse(**user)
        return {"user": user_response, "message": "Login successful"}
        
    except HTTPException:
        # Re-raise HTTP exceptions (like invalid credentials)
        raise
    except Exception as e:
        logging.error(f"Unexpected error during login for user {user_data.username}: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")

@api_router.post("/auth/logout")
async def logout_user(response: Response, user: dict = Depends(require_auth)):
    # Delete session from database
    await db.sessions.delete_many({"user_id": user["id"]})
    
    # Clear cookie
    response.delete_cookie(key="session_token", path="/")
    
    return {"message": "Logout successful"}

@api_router.get("/auth/me")
async def get_current_user_info(user: dict = Depends(require_auth)):
    return UserResponse(**user)

@api_router.post("/auth/onboarding")
async def complete_onboarding(onboarding_data: OnboardingData, user: dict = Depends(require_auth)):
    """Complete user onboarding by setting their name"""
    # Update user with name and mark onboarding as completed
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {
            "name": onboarding_data.name,
            "onboarding_completed": True
        }}
    )
    
    # Get updated user
    updated_user = await db.users.find_one({"id": user["id"]})
    updated_user = parse_from_mongo(updated_user)
    
    return {"message": "Onboarding completed successfully", "user": UserResponse(**updated_user)}

# Google OAuth Routes
@api_router.get("/auth/google")
async def google_auth():
    redirect_url = "https://hybrid-chat-app.preview.emergentagent.com/dashboard"
    google_auth_url = f"https://auth.emergentagent.com/?redirect={redirect_url}"
    return {"auth_url": google_auth_url}

@api_router.post("/auth/google/callback")
async def google_callback(request: Request, response: Response):
    # Get session_id from request
    body = await request.json()
    session_id = body.get("session_id")
    
    if not session_id:
        raise HTTPException(status_code=400, detail="Missing session_id")
    
    try:
        # Get user data from Emergent Auth
        async with httpx.AsyncClient() as client:
            auth_response = await client.get(
                "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data",
                headers={"X-Session-ID": session_id}
            )
            
            if auth_response.status_code != 200:
                raise HTTPException(status_code=400, detail="Invalid session")
            
            user_data = auth_response.json()
            
        # Check if user exists
        existing_user = await db.users.find_one({"email": user_data["email"]})
        
        if existing_user:
            user = parse_from_mongo(existing_user)
        else:
            # Create new user from Google data
            user = User(
                username=user_data["email"].split("@")[0],  # Use email prefix as username
                email=user_data["email"],
                password_hash="",  # No password for OAuth users
                name=user_data.get("name", ""),  # Get name from Google if available
                is_verified=True,  # Google users are auto-verified
                onboarding_completed=bool(user_data.get("name")),  # Complete onboarding if name exists
                oauth_provider="google",
                oauth_id=user_data["id"]
            )
            
            user_dict = prepare_for_mongo(user.dict())
            await db.users.insert_one(user_dict)
            user = user.dict()
        
        # Create session token
        session_token = user_data["session_token"]
        expires_at = datetime.now(timezone.utc) + timedelta(days=7)
        
        # Save session to database
        session = Session(
            user_id=user["id"],
            session_token=session_token,
            expires_at=expires_at
        )
        session_dict = prepare_for_mongo(session.dict())
        await db.sessions.insert_one(session_dict)
        
        # Update last login
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"last_login": datetime.now(timezone.utc).isoformat()}}
        )
        
        # Set cookie
        response.set_cookie(
            key="session_token",
            value=session_token,
            httponly=True,
            secure=True,
            samesite="none",
            path="/",
            expires=expires_at
        )
        
        user_response = UserResponse(**user)
        return {"user": user_response, "message": "Google login successful"}
        
    except Exception as e:
        logging.error(f"Google OAuth error: {e}")
        raise HTTPException(status_code=400, detail="OAuth authentication failed")

# Admin Routes
@api_router.get("/admin/stats", response_model=AdminStats)
async def get_admin_stats(admin: dict = Depends(require_admin)):
    # Get statistics
    total_users = await db.users.count_documents({})
    verified_users = await db.users.count_documents({"is_verified": True})
    total_conversations = await db.conversations.count_documents({})
    total_messages = await db.messages.count_documents({})
    
    # Get recent users
    recent_users_data = await db.users.find().sort("created_at", -1).limit(10).to_list(10)
    recent_users = [UserResponse(**parse_from_mongo(user)) for user in recent_users_data]
    
    return AdminStats(
        total_users=total_users,
        verified_users=verified_users,
        total_conversations=total_conversations,
        total_messages=total_messages,
        recent_users=recent_users
    )

@api_router.get("/admin/users", response_model=List[UserResponse])
async def get_all_users(admin: dict = Depends(require_admin)):
    users = await db.users.find().sort("created_at", -1).to_list(1000)
    return [UserResponse(**parse_from_mongo(user)) for user in users]

# Debug Routes
@api_router.get("/debug/info")
async def debug_info():
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}

@api_router.post("/debug/test-vision")
async def test_vision_api(image_data: dict):
    """Test Vision API directly"""
    try:
        # This is just a test - normally we'd get from uploaded file
        # test_question = image_data.get("question", "Bu resimde ne var?")
        
        # Try with a simple test
        headers = {
            "Authorization": f"Bearer {EMERGENT_LLM_KEY}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "gpt-4o-mini", 
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "Test message for Vision API"}
                    ]
                }
            ],
            "max_tokens": 100
        }
        
        async with httpx.AsyncClient() as client:
            response = await client.post(
                "https://api.openai.com/v1/chat/completions",
                headers=headers,
                json=payload,
                timeout=30.0
            )
            
            return {
                "status_code": response.status_code,
                "response_text": response.text[:500],
                "headers": dict(response.headers),
                "emergent_key_prefix": EMERGENT_LLM_KEY[:10] + "..." if EMERGENT_LLM_KEY else "None"
            }
    except Exception as e:
        return {"error": str(e), "type": str(type(e))}

# Chat Routes (Protected)
# Anonymous user - no auth required
ANONYMOUS_USER_ID = "anonymous"

@api_router.get("/conversations", response_model=List[Conversation])
async def get_conversations():
    conversations = await db.conversations.find({"user_id": ANONYMOUS_USER_ID}).sort("updated_at", -1).to_list(1000)
    return [Conversation(**parse_from_mongo(conv)) for conv in conversations]

@api_router.post("/conversations", response_model=Conversation)
async def create_conversation(input: ConversationCreate):
    conversation = Conversation(user_id=ANONYMOUS_USER_ID, **input.dict())
    conversation_dict = prepare_for_mongo(conversation.dict())
    await db.conversations.insert_one(conversation_dict)
    return conversation

@api_router.get("/conversations/{conversation_id}/messages", response_model=List[MessageResponse])
async def get_messages(conversation_id: str):
    # Check if conversation exists for anonymous user
    conversation = await db.conversations.find_one({"id": conversation_id, "user_id": ANONYMOUS_USER_ID})
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversation not found")
    
    messages = await db.messages.find({"conversation_id": conversation_id}).sort("timestamp", 1).to_list(1000)
    return [MessageResponse(**parse_from_mongo(msg)) for msg in messages]

@api_router.post("/conversations/{conversation_id}/messages", response_model=MessageResponse)
async def send_message(conversation_id: str, input: MessageCreate):
    # Check if conversation exists for anonymous user
    conversation = await db.conversations.find_one({"id": conversation_id, "user_id": ANONYMOUS_USER_ID})
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversation not found")
    
    # Check if this is the first message BEFORE saving it
    existing_message_count = await db.messages.count_documents({"conversation_id": conversation_id})
    is_first_message = existing_message_count == 0
    
    # Save user message
    user_message = Message(
        conversation_id=conversation_id,
        role="user",
        content=input.content
    )
    user_message_dict = prepare_for_mongo(user_message.dict())
    await db.messages.insert_one(user_message_dict)
    
    # Update conversation title if this is the first message
    if is_first_message:
        # Generate meaningful title using the new function
        new_title = generate_conversation_title(input.content)
        logging.info(f"Generated new title for conversation {conversation_id}: {new_title}")
            
        # Update conversation title
        await db.conversations.update_one(
            {"id": conversation_id},
            {"$set": {"title": new_title, "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    try:
        # SMART HYBRID SYSTEM: Quick analysis and intelligent routing
        logging.info(f"Using SMART HYBRID SYSTEM for question: {input.content}")
        
        # Check if the question is about uploaded files
        file_content = None
        file_name = None
        file_path = None
        file_type = None
        
        # Always check for uploaded files in the conversation
        # Get the most recent uploaded file for this conversation
        recent_file = await db.file_uploads.find_one(
            {"conversation_id": conversation_id},
            sort=[("uploaded_at", -1)]
        )
        
        if recent_file and os.path.exists(recent_file["file_path"]):
            file_path = recent_file["file_path"]
            file_name = recent_file["file_name"]
            file_type = recent_file["file_type"]
            
            # For images, always use ChatGPT Vision when there's an uploaded image
            if file_type in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp']:
                logging.info(f"Uploaded image detected, using ChatGPT Vision: {file_name}")
                ai_content = await process_image_with_chatgpt_vision(input.content, file_path, file_name)
                processed = True
            else:
                # Extract text from non-image files and include context
                file_content = await extract_text_from_file(file_path, file_type)
                logging.info(f"Uploaded file detected, using file for context: {file_name}")
                processed = False
        else:
            logging.info("No uploaded file found in conversation")
            processed = False
        
        # Only process with hybrid system if not already processed (e.g., not an image)
        if not processed:
            # Check version and route accordingly
            if input.version == "free":
                logging.info("FREE version selected - using Ollama AnythingLLM")
                ai_content = await process_with_ollama_free(input.content, input.conversationMode, file_content, file_name)
            else:
                # PRO version - use simple system: Current topics → Web Search, Others → AnythingLLM → GPT-5-nano
                logging.info("PRO version selected - using simple system")
                ai_content = await simple_pro_system(input.content, input.conversationMode, file_content, file_name)
        
            logging.info("AI processing completed successfully")
                
    except Exception as e:
        logging.error(f"Smart hybrid system error: {e}")
        ai_content = "Üzgünüm, şu anda teknik bir sorun yaşıyorum. Lütfen sorunuzu tekrar deneyin."
    
    # Save AI response
    ai_message = Message(
        conversation_id=conversation_id,
        role="assistant",
        content=ai_content
    )
    ai_message_dict = prepare_for_mongo(ai_message.dict())
    await db.messages.insert_one(ai_message_dict)
    
    # Update conversation timestamp
    await db.conversations.update_one(
        {"id": conversation_id},
        {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return MessageResponse(**ai_message.dict())

@api_router.delete("/conversations/{conversation_id}")
async def delete_conversation(conversation_id: str):
    # Check if conversation exists for anonymous user
    conversation = await db.conversations.find_one({"id": conversation_id, "user_id": ANONYMOUS_USER_ID})
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversation not found")
    
    # Delete conversation and associated messages
    await db.conversations.delete_one({"id": conversation_id})
    await db.messages.delete_many({"conversation_id": conversation_id})
    
    # Delete associated uploaded files
    await db.file_uploads.delete_many({"conversation_id": conversation_id})
    
    return {"message": "Conversation deleted successfully"}

@api_router.post("/conversations/{conversation_id}/upload")
async def upload_file(
    conversation_id: str,
    file: UploadFile = File(...),
):
    # Check if conversation exists for anonymous user
    conversation = await db.conversations.find_one({"id": conversation_id, "user_id": ANONYMOUS_USER_ID})
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversation not found")
    
    # Check file size
    if file.size > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {MAX_FILE_SIZE / (1024*1024):.1f}MB")
    
    # Check file type
    file_type = get_file_type(file.filename)
    allowed_types = ['pdf', 'xlsx', 'xls', 'docx', 'txt', 'jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp']
    if file_type not in allowed_types:
        raise HTTPException(status_code=400, detail=f"Unsupported file type. Allowed: {', '.join(allowed_types)}")
    
    try:
        # Create unique filename
        unique_filename = f"{uuid4()}_{file.filename}"
        file_path = UPLOAD_DIR / unique_filename
        
        # Save file to disk
        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Save file info to database
        file_upload = FileUpload(
            conversation_id=conversation_id,
            file_name=file.filename,
            file_type=file_type,
            file_path=str(file_path)
        )
        
        file_dict = prepare_for_mongo(file_upload.dict())
        await db.file_uploads.insert_one(file_dict)
        
        # Extract text from file
        extracted_text = await extract_text_from_file(str(file_path), file_type)
        
        # Auto-generate a system message about the uploaded file
        if file_type in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp']:
            file_icon = "🖼️"
            file_description = "Bu görsel hakkında soru sorabilir, içeriğini analiz ettirebilir, görseldeki metni okutabilir veya açıklama isteyebilirsiniz."
        else:
            file_icon = "📎"
            file_description = "Bu dosya hakkında soru sorabilir, özet çıkartabilir, çeviri yaptırabilir veya analiz edebilirsiniz."
        
        system_message = Message(
            conversation_id=conversation_id,
            role="assistant",
            content=f"{file_icon} **{file.filename}** dosyası başarıyla yüklendi!\n\nDosya türü: {file_type.upper()}\nDosya boyutu: {file.size / 1024:.1f} KB\n\n{file_description}"
        )
        
        system_message_dict = prepare_for_mongo(system_message.dict())
        await db.messages.insert_one(system_message_dict)
        
        # Update conversation timestamp
        await db.conversations.update_one(
            {"id": conversation_id},
            {"$set": {"updated_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {
            "message": "File uploaded successfully",
            "file_id": file_upload.id,
            "file_name": file.filename,
            "file_type": file_type,
            "system_message": system_message.dict()
        }
        
    except Exception as e:
        logging.error(f"File upload error: {e}")
        # Clean up file if it was created
        if 'file_path' in locals() and os.path.exists(file_path):
            os.remove(file_path)
        raise HTTPException(status_code=500, detail=f"File upload failed: {str(e)}")

@api_router.get("/conversations/{conversation_id}/files")
async def get_uploaded_files(conversation_id: str):
    # Check if conversation exists for anonymous user
    conversation = await db.conversations.find_one({"id": conversation_id, "user_id": ANONYMOUS_USER_ID})
    if not conversation:
        raise HTTPException(status_code=404, detail="Conversation not found")
    
    # Get uploaded files for this conversation
    files = await db.file_uploads.find({"conversation_id": conversation_id}).sort("uploaded_at", -1).to_list(100)
    
    return [
        {
            "id": file["id"],
            "file_name": file["file_name"],
            "file_type": file["file_type"],
            "uploaded_at": file["uploaded_at"]
        }
        for file in files
    ]

# Report endpoints
@api_router.post("/reports", response_model=ReportResponse)
async def create_report(input: ReportCreate, user: dict = Depends(require_auth)):
    """Create a new bug report"""
    report = Report(
        user_id=user["id"],
        message=input.message,
        user_agent=input.user_agent,
        url=input.url
    )
    
    report_dict = prepare_for_mongo(report.dict())
    await db.reports.insert_one(report_dict)
    
    return ReportResponse(**report.dict())

@api_router.get("/admin/reports", response_model=List[ReportResponse])
async def get_reports(user: dict = Depends(require_admin)):
    """Get all bug reports (admin only)"""
    reports = await db.reports.find().sort("created_at", -1).to_list(length=None)
    return [ReportResponse(**parse_from_mongo(report)) for report in reports]

@api_router.patch("/admin/reports/{report_id}/status")
async def update_report_status(report_id: str, status: str, user: dict = Depends(require_admin)):
    """Update report status (admin only)"""
    if status not in ["open", "investigating", "resolved"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    result = await db.reports.update_one(
        {"id": report_id},
        {"$set": {"status": status}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Report not found")
    
    return {"message": "Report status updated successfully"}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("startup")
async def startup_event():
    await init_admin()

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()